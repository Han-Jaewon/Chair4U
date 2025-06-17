import os, shutil, string, re
from itertools import product
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# ─────────────────────────── 경  로 ───────────────────────────
EXCEL_FOLDER       = r"C:\Users\grace\OneDrive\Desktop\dataset\중복 확인 필요\엑셀"
IMG_FOLDER_MALE    = r"C:\Users\grace\OneDrive\Desktop\dataset\중복 확인 필요\이미지\남성"
IMG_FOLDER_FEMALE  = r"C:\Users\grace\OneDrive\Desktop\dataset\중복 확인 필요\이미지\여성"
LOG_DIR            = r"C:\Users\grace\OneDrive\Desktop\dataset\엑셀\이름 변경 로그"

ALIAS_EXCEL_FOLDER = os.path.join(EXCEL_FOLDER, "가명")
ALIAS_IMG_MALE     = os.path.join(IMG_FOLDER_MALE , "[가명]")
ALIAS_IMG_FEM      = os.path.join(IMG_FOLDER_FEMALE, "[가명]")
for p in (ALIAS_EXCEL_FOLDER, ALIAS_IMG_MALE, ALIAS_IMG_FEM, LOG_DIR):
    os.makedirs(p, exist_ok=True)

# ─────────────────────────── 헬퍼 ───────────────────────────
def read_visible_rows(xlsx):
    wb = load_workbook(xlsx, data_only=True)
    return pd.DataFrame(
        [[c.value for c in r[:4]]
         for r in wb.active.iter_rows(min_row=2)
         if not wb.active.row_dimensions[r[0].row].hidden],
        columns=["이름","키","수집링크","성별"]
    )

def nrm_name(name:str)->str:                # 이름 정규화
    s = str(name).strip().lower()
    s = re.sub(r'[ -]+', '_', s)
    s = re.sub(r'[^a-z0-9_가-힣]', '', s)
    return re.sub(r'__+', '_', s)

def token_key(nm):                          # 성·이름 순서 무시용
    return frozenset(nm.split('_'))

def nrm_h(h):
    try: return int(float(h))
    except: return str(h).strip()

def alias_sort_key(a):                      # a,b,…,z,aa,ab 정렬
    return [string.ascii_lowercase.index(c) for c in a]

def past_aliases(log_dir):
    """이전 모든 로그에서 이미 발급된 가명 집합 반환"""
    al=set()
    for f in os.listdir(log_dir):
        if f.startswith("anonymization_log_") and f.endswith(".xlsx"):
            try:
                df=pd.read_excel(os.path.join(log_dir,f), usecols=["가명"])
                al.update(df["가명"].dropna().astype(str))
            except: pass
    return al

def alias_gen(start_after=None):
    """start_after 다음 값부터 무한 생성"""
    passed = start_after is None
    length=1
    while True:
        for combo in product(string.ascii_lowercase, repeat=length):
            al=''.join(combo)
            if not passed:
                passed = (al==start_after)
                continue
            yield al
        length+=1

def split_fname(fname):
    """[엑셀]_[이름](_seq)_키.ext → (excel,raw,seq,height,ext)"""
    base,ext=os.path.splitext(fname)
    pr=base.split('_')
    if len(pr)<3 or not pr[-1].isdigit(): return None
    excel,height=pr[0],int(pr[-1])
    if pr[-2].isdigit(): seq=pr[-2]; raw='_'.join(pr[1:-2])
    else:                   seq=None ; raw='_'.join(pr[1:-1])
    return excel,raw,seq,height,ext.lstrip('.')

EXTS=(".jpg",".jpeg",".png",".bmp",".gif",".webp")

# ──────────────────── 가명 매핑 초기화 ────────────────────
used_alias      = past_aliases(LOG_DIR)
last_alias      = max(used_alias, key=alias_sort_key) if used_alias else None
alias_iter      = alias_gen(last_alias)
map_exact       = {}   # (name_norm,height)      -> alias
map_tokens      = {}   # (token_set,height)      -> alias

logs=[]
def wlog(act,nm,al,fp,ex=""):
    logs.append(dict(시각=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     실명=nm,가명=al,작업=act,파일=fp,추가=ex))

# ────────────────────── 엑셀 처리 ──────────────────────
print("📑 새 엑셀 처리")
for xf in os.listdir(EXCEL_FOLDER):
    if not xf.endswith(".xlsx") or xf.startswith(("~$","anonymization_log_")): continue
    path=os.path.join(EXCEL_FOLDER,xf)
    df=read_visible_rows(path)
    if df.empty: continue
    df["실제이름"]=None
    out=df.copy()

    for i,r in df.iterrows():
        name, h = r["이름"], r["키"]
        if pd.isna(name): continue
        n_norm=nrm_name(name); h_n=nrm_h(h)
        k_exact=(n_norm,h_n); k_tok=(token_key(n_norm),h_n)

        al = map_exact.get(k_exact) or map_tokens.get(k_tok)
        if al is None:
            al=next(alias_iter)
            while al in used_alias: al=next(alias_iter)
            used_alias.add(al)
            map_exact[k_exact]=al; map_tokens[k_tok]=al
            wlog("ALIAS_NEW",name,al,path,f"키:{h}")
        out.at[i,"실제이름"]=name; out.at[i,"이름"]=al
        wlog("NAME_REPLACED",name,al,path)

    out.to_excel(os.path.join(ALIAS_EXCEL_FOLDER,xf),index=False)
    print("  •",xf,"→ 가명본 저장")

# ───────────────────── 이미지 처리 ─────────────────────
unmatched=[]
def proc_imgs(src,dst,gender):
    rn=cp=0
    for f in os.listdir(src):
        if not f.lower().endswith(EXTS): continue
        info=split_fname(f)
        if not info:
            shutil.copy2(os.path.join(src,f),os.path.join(dst,f));cp+=1;continue
        excel,raw,seq,h,ext=info
        n_norm=nrm_name(raw); h_n=nrm_h(h)
        al = map_exact.get((n_norm,h_n)) or map_tokens.get((token_key(n_norm),h_n))
        if al is None:
            unmatched.append(f"{gender}:{f}")
            shutil.copy2(os.path.join(src,f),os.path.join(dst,f));cp+=1;continue
        newf=f"{excel}_{al}_{seq+'_' if seq else ''}{h}.{ext}"
        shutil.copy2(os.path.join(src,f),os.path.join(dst,newf));rn+=1
        wlog("IMG_RENAMED",raw,al,os.path.join(dst,newf))
    print(f"{gender}: rename {rn}, copy {cp}")

print("\n 이미지 복사·치환")
proc_imgs(IMG_FOLDER_MALE , ALIAS_IMG_MALE,"남성")
proc_imgs(IMG_FOLDER_FEMALE, ALIAS_IMG_FEM ,"여성")

# ──────────────────── 로그 & 미매칭 ────────────────────
log_file=os.path.join(LOG_DIR,"anonymization_log_"+datetime.now().strftime("%Y%m%d_%H%M%S")+".xlsx")
pd.DataFrame(logs).to_excel(log_file,index=False); print("\n📋 로그 저장:",log_file)

if unmatched:
    miss=os.path.join(LOG_DIR,"미매칭_이미지.txt")
    with open(miss,"w",encoding="utf-8") as fp: fp.write("\n".join(unmatched))
    print("⚠️ 매칭 실패",len(unmatched),"건 >",miss)

print("\n 새 데이터 가명 처리 완료")
